from tkinter import W
from openpyxl import Workbook, load_workbook

# 加载一个现有的Excel
book = load_workbook("sample.xlsx")
sheet1 = book.active

'''
读取单个单元格
'''
a1 = sheet1['A1']
a2 = sheet1.cell(row=3,column=1)

print(a1.value, a2.value)


'''
读取多个单元格
'''

cells = sheet1['A1':'B10']

for c1, c2 in cells:
    # print("{0:8} {1:8}".format(c1.value, c2.value))
    print(c1.value,  c2.value)


'''
Openpyxl 按行迭代
iter_rows()方法将工作表中的单元格返回为行。
'''